# Import the modules needed to run the script.
from termcolor import colored
import sys
import os
from openpyxl import load_workbook
wb = load_workbook("C:\\Users\\plesn\\Documents\\Programowanie\\pythonprojects\\untitled\\oceny.xlsx")
sheetnames = wb.sheetnames

# Main definition - constants
menu_actions = {}


# =======================
#     MENU
# =======================

# Main menu
def main_menu():

    # os.system('cls')

    print(colored("\u001b[1m Witaj w systemie", 'blue'),colored("\u001b[1m Dziennik Ocen,\n","magenta"))
    print(colored("\u001b[4m Wybierz operacje z menu:","blue"))
    print(colored("[1]",'blue'), "Średnia ocen z przedmiotu")
    print(colored("[2]",'blue'), "Średnia ocen dla studenta")
    print(colored("[3]",'blue'), "Średnia ocen ze wszystkich przedmiotów")
    print(colored("[4]",'blue'), "Ranking studentów z przedmiotu")
    print(colored("[5]",'blue'), "Ranking studentów dla wsystkich przedmiotów")
    print(colored("[6]",'blue'), "Zakończ dzialanie programu")
    choice = input(colored("\u001b[1m Twój wybór >>  ","red"))
    exec_menu(choice)

    return


# Execute menu
def exec_menu(choice):
    # os.system('clear')
    ch = choice.lower()
    if ch == '':
        menu_actions['main_menu']()
    elif ch == '1':
        subject = input(colored("Podaj nazwę przedmiotu >>  ","red"))
        calculateAverageForSubject(subject)
        koniec=input(colored("\u001b[1m Kończymy? [t/n]>>","red"))
        if koniec=="t":
            sys.exit()
        else:
            print("\n")
            menu_actions['main_menu']()
    elif ch == '2':
        student = input(colored("Podaj dane studenta >>  ","red"))
        calculateAverageForStudent(student)
        koniec = input(colored("\u001b[1m Kończymy? [t/n]>>", "red"))
        if koniec == "t":
            sys.exit()
        else:
            print("\n")
            menu_actions['main_menu']()
    elif ch == '3':
        calculateAverageForAllSubjects()
        koniec = input(colored("\u001b[1m Kończymy? [t/n]>>", "red"))
        if koniec == "t":
            sys.exit()
        else:
            print("\n")
            menu_actions['main_menu']()
    elif ch == '4':
        subject = input(colored("Podaj nazwę przedmiotu >>  ","red"))
        my_set = set(sheetnames)
        if subject in my_set:
            print("Ranking studentów dla przedmiotu %s" % subject)
            generateStudentRankingForSubject(subject)
        else:
            print("Nie ma takiego przedmiotu jak %s" % subject)
        koniec = input(colored("\u001b[1m Kończymy? [t/n]>>", "red"))
        if koniec == "t":
            sys.exit()
        else:
            print("\n")
            menu_actions['main_menu']()
    elif ch == '5':
        print("Ranking studentów dla wszystkich przedmiotów: %s" % sheetnames)
        generateStudentRankingForAllSubjects()
        koniec = input(colored("\u001b[1m Kończymy? [t/n]>>", "red"))
        if koniec == "t":
            sys.exit()
        else:
            print("\n")
            menu_actions['main_menu']()
    else:
        try:
            menu_actions[ch]()
        except KeyError:
            print("Operacja spoza zakresu, wybierz ponownie operacje z menu [0-6].\n")
            print("\n")
            menu_actions['main_menu']()
    return
#######################################################################
def calculateAverageForSubject(subject):
    points=0
    my_set=set(sheetnames)
    if subject in my_set:
        sheet = wb[subject]
        for col in sheet.iter_cols(min_col=2,max_col=2):
            for cell in col:
                num = len(col)
                points += cell.value
                avg = round(points / num,2)
            return print('Średnia ocen dla przedmiotu %s wynosi %0.2f \033[1;31m' % (subject,avg))
    else:
        return print("Nie ma takiego przedmiotu jak %s" % subject)

######################################################################
def calculateAverageForStudent(student):
    n = len(sheetnames)
    sheet = wb.active
    result = []

    for col in sheet.iter_cols(min_col=1, max_col=1):
        for cell in col:
            result.append(cell.value)

    res_set = set(result)

    if student in res_set:
        points = 0
        for sheet in wb:
            for row in sheet.iter_rows():
                if (row[0].value== student):
                    points = float(row[1].value) + points
                    avgg=round(points/n,2)

        return print('Średnia ocena studenta %s w przedmiotach %s wynosi %0.2f ' %(student,sheetnames,avgg))
    else:
        print("Student %s nie został znaleziony." % student)
#####################################################################

# calculateAverageForAllSubjects
def calculateAverageForAllSubjects():
    for col in wb.active.iter_cols(min_col=1, max_col=1):
        n = len(sheetnames)
        points2_list=[]
        for cell in col:
            points2 = 0
            num = len(col)
            for sheet in wb:
                for row in sheet.iter_rows():
                    if (row[0].value == cell.value):
                        points2 = float(row[1].value) + points2
            points2_list.append(points2)
        return print("Średnia wszystkich ocen wynosi %0.2f" % round(sum(points2_list)/(n*num),2))
###############################################

def generateStudentRankingForSubject(subject):
    students = {}

    sheet=wb[subject]
    for row in sheet:
        student = row[0].value
        points = row[1].value
        if student in students:
            students[student] = points + students[student]
        else:
            students[student] = points

    students_list = []
    for student, points in students.items():
        students_list.append((points, student))

    students_list.sort(reverse=True)

    sorted_students = []
    for points, student in students_list:
        sorted_students.append((student, points))

    for student, points in sorted_students:
        print(student, ':', points)


#######################################################
def generateStudentRankingForAllSubjects():
    for col in wb.active.iter_cols(min_col=1, max_col=1):
        n = len(sheetnames)
        students_list = []
        for cell in col:
            points = 0
            for sheet in wb:
                for row in sheet.iter_rows():
                    if (row[0].value == cell.value):
                        points = float(row[1].value) + points
                    avg=points/n
            students_list.append((avg,cell.value))
        students_list.sort(reverse=True)
        sorted_students = []
        for avg, cell.value in students_list:
            sorted_students.append((cell.value, avg))
        for cell.value, avg in sorted_students:
             print(cell.value, ':', round(avg,2))
        return


# Back to main menu
def back():
    menu_actions['main_menu']()


# Exit program
def exit():
    sys.exit()


# =======================
#    MENUS DEFINITIONS
# =======================

# Menu definition
menu_actions = {
    'main_menu': main_menu,
    '1': calculateAverageForSubject,
    '2': calculateAverageForStudent,
    '3': calculateAverageForAllSubjects,
    '4': generateStudentRankingForSubject,
    '5': generateStudentRankingForAllSubjects,
    '9': back,
    '6': exit,
}

# =======================
#      MAIN PROGRAM
# =======================

# Main Program
if __name__ == "__main__":
    # Launch main menu

    main_menu()
